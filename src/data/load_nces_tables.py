from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook


MARKER_VALUES = {"#", "!", "‡"}


@dataclass
class ColumnSpec:
    column_group: str
    column_label: str
    value_col: int
    flag_col: int | None


@dataclass
class TableCell:
    table_id: str
    table_title: str
    section_label: str
    row_label: str
    row_number: int
    column_group: str
    column_label: str
    value: float | None
    value_flag: str | None
    source_file: str
    value_type: str


def list_estimate_files(data_dir: str | Path) -> list[Path]:
    base = Path(data_dir)
    return sorted(
        path for path in base.glob("characteristicsofpostsecondary_*.xlsx")
        if not path.stem.endswith("se")
    )


def list_workbook_files(data_dir: str | Path) -> list[Path]:
    return sorted(Path(data_dir).glob("characteristicsofpostsecondary_*.xlsx"))


def paired_files(data_dir: str | Path) -> list[tuple[Path, Path | None]]:
    pairs: list[tuple[Path, Path | None]] = []
    for estimate_path in list_estimate_files(data_dir):
        se_path = estimate_path.with_name(f"{estimate_path.stem}se.xlsx")
        pairs.append((estimate_path, se_path if se_path.exists() else None))
    return pairs


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    return " ".join(text.split())


def _extract_table_id(title_cell: str, fallback: str) -> str:
    match = re.search(r"table\s+(\d+)", title_cell, flags=re.IGNORECASE)
    if match:
        return match.group(1)
    fallback_match = re.search(r"_(\d+)(?:se)?$", fallback, flags=re.IGNORECASE)
    if fallback_match:
        return fallback_match.group(1)
    cleaned = title_cell.replace("Standard Errors for", "").replace("Standard errors for", "").replace("Table", "").replace(".", "").strip()
    return cleaned or fallback


def _is_numeric_text(text: str) -> bool:
    if not text:
        return False
    try:
        float(text)
    except ValueError:
        return False
    return True


def _is_measure_token(text: str) -> bool:
    return _is_numeric_text(text) or text in MARKER_VALUES


def _read_sheet_rows(path: str | Path) -> tuple[list[list[object]], dict[tuple[int, int], str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[list[object]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))

    merged_text: dict[tuple[int, int], str] = {}
    for merged_range in ws.merged_cells.ranges:
        top_left_text = _clean_text(ws.cell(merged_range.min_row, merged_range.min_col).value)
        if not top_left_text:
            continue
        for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
            for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                merged_text[(row_idx, col_idx)] = top_left_text

    return rows, merged_text


def _row_text(rows: list[list[object]], row_idx: int, col_idx: int) -> str:
    if row_idx < 1 or row_idx > len(rows):
        return ""
    row = rows[row_idx - 1]
    if col_idx < 1 or col_idx > len(row):
        return ""
    return _clean_text(row[col_idx - 1])


def _display_text(rows: list[list[object]], merged_text: dict[tuple[int, int], str], row_idx: int, col_idx: int) -> str:
    text = _row_text(rows, row_idx, col_idx)
    if text:
        return text
    return merged_text.get((row_idx, col_idx), "")


def _cell_map(row: list[object]) -> str:
    parts = []
    for col_idx, value in enumerate(row, start=1):
        text = _clean_text(value)
        if text:
            parts.append(f"{col_idx}:{text}")
    return " | ".join(parts)


def _find_first_data_row(rows: list[list[object]]) -> int:
    for row_idx, row in enumerate(rows, start=1):
        row_label = _clean_text(row[0]) if row else ""
        if not row_label:
            continue
        if any(_is_measure_token(_clean_text(value)) for value in row[1:]):
            return row_idx
    raise ValueError("Could not locate first data row")


def _header_rows(rows: list[list[object]], merged_text: dict[tuple[int, int], str], first_data_row: int) -> list[int]:
    max_cols = max((len(row) for row in rows), default=0)
    header_rows: list[int] = []
    for row_idx in range(4, first_data_row):
        if any(_display_text(rows, merged_text, row_idx, col_idx) for col_idx in range(2, max_cols + 1)):
            header_rows.append(row_idx)
    return header_rows


def _header_texts_for_column(
    rows: list[list[object]],
    merged_text: dict[tuple[int, int], str],
    header_rows: list[int],
    col_idx: int,
) -> list[str]:
    texts: list[str] = []
    for row_idx in header_rows:
        text = _display_text(rows, merged_text, row_idx, col_idx)
        if text and (not texts or texts[-1] != text):
            texts.append(text)
    return texts


def _build_column_specs(
    rows: list[list[object]],
    merged_text: dict[tuple[int, int], str],
    first_data_row: int,
) -> list[ColumnSpec]:
    max_cols = max((len(row) for row in rows), default=0)
    header_rows = _header_rows(rows, merged_text, first_data_row)
    measure_cols = [
        col_idx
        for col_idx in range(2, max_cols + 1)
        if _is_measure_token(_row_text(rows, first_data_row, col_idx))
    ]
    measure_set = set(measure_cols)

    specs: list[ColumnSpec] = []
    for col_idx in measure_cols:
        texts = _header_texts_for_column(rows, merged_text, header_rows, col_idx)
        if not texts:
            continue
        previous_texts = _header_texts_for_column(rows, merged_text, header_rows, col_idx - 1) if col_idx > 2 else []
        flag_col = col_idx - 1 if col_idx - 1 not in measure_set and previous_texts == texts else None
        specs.append(
            ColumnSpec(
                column_group=" | ".join(texts[:-1]),
                column_label=texts[-1],
                value_col=col_idx,
                flag_col=flag_col,
            )
        )

    return specs


def _row_has_values(rows: list[list[object]], row_idx: int, column_specs: list[ColumnSpec]) -> bool:
    for spec in column_specs:
        if _row_text(rows, row_idx, spec.value_col):
            return True
        if spec.flag_col is not None and _row_text(rows, row_idx, spec.flag_col):
            return True
    return False


def _find_last_data_row(rows: list[list[object]], first_data_row: int, column_specs: list[ColumnSpec]) -> int:
    last_data_row = first_data_row
    for row_idx in range(first_data_row, len(rows) + 1):
        if _row_has_values(rows, row_idx, column_specs):
            last_data_row = row_idx
    return last_data_row


def _value_and_flag(value_text: str, flag_text: str) -> tuple[float | None, str | None]:
    value = float(value_text) if _is_numeric_text(value_text) else None
    flags: list[str] = []
    for text in (flag_text, value_text):
        if text and not _is_numeric_text(text):
            flags.append(text)
    deduped_flags = list(dict.fromkeys(flags))
    flag = " ".join(deduped_flags) if deduped_flags else None
    return value, flag


def extract_table(path: str | Path, value_type: str = "estimate") -> pd.DataFrame:
    rows, merged_text = _read_sheet_rows(path)
    if len(rows) < 2:
        raise ValueError(f"Unexpected short table: {path}")

    title_cell = _row_text(rows, 2, 1)
    subtitle_cell = _row_text(rows, 2, 2)
    table_title = subtitle_cell or title_cell
    table_id = _extract_table_id(title_cell, Path(path).stem)

    first_data_row = _find_first_data_row(rows)
    column_specs = _build_column_specs(rows, merged_text, first_data_row)
    if not column_specs:
        return pd.DataFrame(
            columns=[
                "table_id",
                "table_title",
                "section_label",
                "row_label",
                "row_number",
                "column_group",
                "column_label",
                "value",
                "value_flag",
                "source_file",
                "value_type",
            ]
        )

    last_data_row = _find_last_data_row(rows, first_data_row, column_specs)
    records: list[TableCell] = []
    current_section = ""

    for row_idx in range(first_data_row, last_data_row + 1):
        row_label = _row_text(rows, row_idx, 1)
        if not row_label:
            continue

        if not _row_has_values(rows, row_idx, column_specs):
            current_section = row_label
            continue

        for spec in column_specs:
            value_text = _row_text(rows, row_idx, spec.value_col)
            flag_text = _row_text(rows, row_idx, spec.flag_col) if spec.flag_col is not None else ""
            if not value_text and not flag_text:
                continue
            value, value_flag = _value_and_flag(value_text, flag_text)
            records.append(
                TableCell(
                    table_id=table_id,
                    table_title=table_title,
                    section_label=current_section,
                    row_label=row_label,
                    row_number=row_idx,
                    column_group=spec.column_group,
                    column_label=spec.column_label,
                    value=value,
                    value_flag=value_flag,
                    source_file=Path(path).name,
                    value_type=value_type,
                )
            )

    return pd.DataFrame(record.__dict__ for record in records)


def merge_estimates_and_se(estimate_df: pd.DataFrame, se_df: pd.DataFrame | None) -> pd.DataFrame:
    estimate_subset = estimate_df.rename(columns={"value": "estimate", "value_flag": "estimate_flag"}).drop(
        columns=["value_type"]
    )

    if se_df is None or se_df.empty:
        result = estimate_subset.copy()
        result["standard_error"] = pd.NA
        result["standard_error_flag"] = pd.NA
        return result

    se_subset = se_df.rename(columns={"value": "standard_error", "value_flag": "standard_error_flag"}).drop(
        columns=["value_type", "source_file"]
    )

    merge_keys = [
        "table_id",
        "table_title",
        "section_label",
        "row_label",
        "row_number",
        "column_group",
        "column_label",
    ]
    return estimate_subset.merge(se_subset, on=merge_keys, how="left")


def load_all_tables(data_dir: str | Path) -> pd.DataFrame:
    merged_tables: list[pd.DataFrame] = []
    for estimate_path, se_path in paired_files(data_dir):
        estimate_df = extract_table(estimate_path, value_type="estimate")
        se_df = extract_table(se_path, value_type="standard_error") if se_path else None
        merged_tables.append(merge_estimates_and_se(estimate_df, se_df))

    if not merged_tables:
        return pd.DataFrame(
            columns=[
                "table_id",
                "table_title",
                "section_label",
                "row_label",
                "row_number",
                "column_group",
                "column_label",
                "source_file",
                "estimate",
                "estimate_flag",
                "standard_error",
                "standard_error_flag",
            ]
        )

    return pd.concat(merged_tables, ignore_index=True)


def scan_table_rows(path: str | Path) -> pd.DataFrame:
    rows, merged_text = _read_sheet_rows(path)
    if len(rows) < 2:
        raise ValueError(f"Unexpected short table: {path}")

    title_cell = _row_text(rows, 2, 1)
    subtitle_cell = _row_text(rows, 2, 2)
    table_title = subtitle_cell or title_cell
    table_id = _extract_table_id(title_cell, Path(path).stem)

    first_data_row = _find_first_data_row(rows)
    column_specs = _build_column_specs(rows, merged_text, first_data_row)
    last_data_row = _find_last_data_row(rows, first_data_row, column_specs) if column_specs else first_data_row

    records: list[dict[str, object]] = []
    current_section = ""
    file_role = "standard_error" if Path(path).stem.endswith("se") else "estimate"

    for row_idx, row in enumerate(rows, start=1):
        cell_map = _cell_map(row)
        if not cell_map:
            continue

        row_label = _row_text(rows, row_idx, 1)
        if row_idx < first_data_row:
            row_type = "header" if row_idx >= 4 else "meta"
            section_label = ""
        elif row_idx <= last_data_row:
            if _row_has_values(rows, row_idx, column_specs):
                row_type = "data"
                section_label = current_section
            else:
                row_type = "section"
                section_label = row_label
                current_section = row_label
        else:
            row_type = "note"
            section_label = ""

        records.append(
            {
                "table_id": table_id,
                "table_title": table_title,
                "source_file": Path(path).name,
                "file_role": file_role,
                "row_number": row_idx,
                "row_type": row_type,
                "section_label": section_label,
                "row_label": row_label,
                "cell_map": cell_map,
            }
        )

    return pd.DataFrame(records)


def scan_all_tables(data_dir: str | Path) -> pd.DataFrame:
    scanned_tables = [scan_table_rows(path) for path in list_workbook_files(data_dir)]
    if not scanned_tables:
        return pd.DataFrame(
            columns=[
                "table_id",
                "table_title",
                "source_file",
                "file_role",
                "row_number",
                "row_type",
                "section_label",
                "row_label",
                "cell_map",
            ]
        )
    return pd.concat(scanned_tables, ignore_index=True)


def filter_rows(df: pd.DataFrame, keywords: Iterable[str]) -> pd.DataFrame:
    pattern = "|".join(keywords)
    return df[df["row_label"].str.contains(pattern, case=False, na=False)].copy()
